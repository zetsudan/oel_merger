from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from starlette.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io, os, re, json, datetime

os.makedirs("static", exist_ok=True)

app = FastAPI(title="oel_merger", description="OEL merge — summary & chart (fix enumerate + interval chart)", version="3.3.0")
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

STEP = 0.0125  # THz
GRID_START = 191.325
GRID_END = 196.125

GRID_POINTS = [round(GRID_START + i*STEP, 5) for i in range(int(round((GRID_END - GRID_START)/STEP)) + 1)]
INTERVAL_LOWER_EDGES = GRID_POINTS[:-1]

STATE = { "oels": [] }  # list of {"name": str, "ranges": str, "free_edges": set[float]}

_range_re = re.compile(r"\(?\s*([0-9]+(?:\.[0-9]+)?)\s*-\s*([0-9]+(?:\.[0-9]+)?)\s*\)?")

def frange_inclusive(start: float, end: float, step: float):
    if start > end:
        start, end = end, start
    def r5(x): return round(x + 0.0, 5)
    def snap(x): return r5(round(x / step) * step)
    s = snap(start); e = snap(end)
    if s < start: s = r5(s + step)
    if e > end: e = r5(e - step)
    cur = s; out = []
    while cur <= e + 1e-9:
        out.append(r5(cur))
        cur = r5(cur + step)
    return out

def parse_ranges_to_edges_set(ranges_text: str) -> set[float]:
    edges = set()
    for seg in ranges_text.split(":"):
        seg = seg.strip()
        if not seg: continue
        m = _range_re.search(seg)
        if not m:
            try:
                v = round(float(seg.strip("() ").replace(",", ".")), 5)
                edges.add(v)
                continue
            except Exception:
                continue
        a = float(m.group(1)); b = float(m.group(2))
        for f in frange_inclusive(a, b, STEP):
            if GRID_START - 1e-9 <= f <= GRID_END + 1e-9:
                edges.add(round(f, 5))
    return edges

def interval_status_for_oel(lower_edge: float, oel) -> str:
    ue = round(lower_edge + STEP, 5)
    return "FREE" if (round(lower_edge,5) in oel["free_edges"] and ue in oel["free_edges"]) else "IN USED"

def precompute_status_matrix():
    rows = []
    summary_mask = []
    for le in INTERVAL_LOWER_EDGES:
        row = []
        all_free = True  # empty => all free
        for o in STATE["oels"]:
            st = interval_status_for_oel(le, o)
            row.append(st)
            if st != "FREE":
                all_free = False
        rows.append(row)
        summary_mask.append(all_free)
    return rows, summary_mask

def context(request, flash=None):
    rows, summary_mask = precompute_status_matrix()
    return {
        "request": request,
        "oels": STATE["oels"],
        "intervals": INTERVAL_LOWER_EDGES,
        "STEP": STEP,
        "oels_len": len(STATE["oels"]),
        "grid_size": len(INTERVAL_LOWER_EDGES),
        "rows": rows,
        "summary_mask": summary_mask,
        "flash": flash,
    }

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", context(request))

@app.get("/summary", response_class=HTMLResponse)
async def summary(request: Request):
    return templates.TemplateResponse("summary.html", context(request))

@app.get("/chart", response_class=HTMLResponse)
async def chart(request: Request):
    lower_edges = INTERVAL_LOWER_EDGES
    upper_edges = [round(le + STEP, 5) for le in lower_edges]
    _, mask = precompute_status_matrix()
    return templates.TemplateResponse("chart.html", {
        "request": request,
        "lower_json": json.dumps(lower_edges),
        "upper_json": json.dumps(upper_edges),
        "mask_json": json.dumps(mask),
        "step": STEP,
    })

@app.get("/add_oel")
async def add_oel_get():
    return RedirectResponse("/", status_code=302)

@app.post("/add_oel", response_class=HTMLResponse)
async def add_oel(request: Request, name: str = Form(""), passband: str = Form("")):
    name = (name or "").strip()
    passband = (passband or "").strip()
    if not name:
        return templates.TemplateResponse("index.html", context(request, {"ok": False, "msg": "Имя OEL не должно быть пустым."}))
    if not passband:
        return templates.TemplateResponse("index.html", context(request, {"ok": False, "msg": "Passband не должен быть пустым."}))
    edges = parse_ranges_to_edges_set(passband)
    STATE["oels"].append({"name": name, "ranges": passband, "free_edges": edges})
    return templates.TemplateResponse("index.html", context(request, {"ok": True, "msg": f"Добавлен OEL: {name}"}))

@app.get("/reset", response_class=HTMLResponse)
async def reset(request: Request):
    STATE["oels"].clear()
    return templates.TemplateResponse("index.html", context(request, {"ok": True, "msg": "Список OEL очищен."}))

@app.get("/download_excel")
async def download_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "oel_merged"

    headers = ["Edge Freq (THz)", "Central Freq (THz)", "Edge Freq (THz)"] + [o["name"] for o in STATE["oels"]] + ["Summary (ALL)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.alignment = Alignment(wrap_text=True)

    fill_green = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    rows, summary_mask = precompute_status_matrix()

    for idx, le in enumerate(INTERVAL_LOWER_EDGES):
        ue = round(le + STEP, 5)
        ce = round(le + STEP/2, 5)
        row = [f"{le:.5f}", f"{ce:.5f}", f"{ue:.5f}"]
        statuses = rows[idx] if STATE["oels"] else []
        for st in statuses:
            row.append(st)
        summary = "FREE" if (summary_mask[idx]) else "IN USED"
        row.append(summary)
        ws.append(row)
        r = ws.max_row
        for col in range(4, 4 + len(statuses)):
            ws.cell(row=r, column=col).fill = fill_green if ws.cell(row=r, column=col).value == "FREE" else fill_red
        ws.cell(row=r, column=4 + len(statuses)).fill = fill_green if summary == "FREE" else fill_red

    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        max_len = 12
        for cell in ws[get_column_letter(col)]:
            v = cell.value
            if v is None: continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 48)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"oel_{ts}.xlsx"
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})
